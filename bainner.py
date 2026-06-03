"""
NFS-e Nacional - Download em Lote de Notas Recebidas (v2)
==========================================================
Empresa: Rezende Construção e Manutenção Ltda
CNPJ Matriz: 28.932.062/0001-50

Novidades da v2:
- Aceita múltiplos certificados .pfx ao mesmo tempo
- Identifica cada um lendo a chave pública (CNPJ, razão social, validade)
- Distingue Matriz vs Filial automaticamente
- Detecta e-CPF e certificados vencidos
- Seletor visual para escolher o certificado correto

Estrutura de saída:
    notas_fiscais_recebidas/
    ├── PDFs/<ano-mes>/<arquivo>.pdf
    ├── XMLs/<ano-mes>/<arquivo>.xml
    └── manifesto.xlsx

Execução:
    streamlit run nfse_recebidas_downloader.py
"""

import base64
import gzip
import io
import re
import tempfile
import time
import unicodedata
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional
from xml.etree import ElementTree as ET

import pandas as pd
import requests
import streamlit as st
from cryptography.hazmat.primitives.serialization import pkcs12
from cryptography.x509.oid import NameOID
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from requests_pkcs12 import Pkcs12Adapter

# ============================================================================
# CONFIGURAÇÃO
# ============================================================================

CNPJ_REZENDE_MATRIZ = "28932062000150"
CNPJ_BASE_REZENDE = "28932062"  # raiz dos 8 primeiros dígitos

ADN_PROD = "https://adn.nfse.gov.br"
ADN_HOMOLOG = "https://adn.producaorestrita.nfse.gov.br"
NFSE_PROD = "https://www.nfse.gov.br"
NFSE_HOMOLOG = "https://www.producaorestrita.nfse.gov.br"

NS = {"nfse": "http://www.sped.fazenda.gov.br/nfse"}

ORANGE = "#F7931E"
BG = "#0A0A0A"

# ============================================================================
# UTILITÁRIOS GERAIS
# ============================================================================

def slugify(texto: str, max_len: int = 30) -> str:
    if not texto:
        return "SEM_NOME"
    texto = unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode()
    texto = re.sub(r"[^A-Za-z0-9]+", "_", texto).strip("_").upper()
    return texto[:max_len] or "SEM_NOME"


def formatar_cnpj(cnpj: str) -> str:
    """28932062000150 -> 28.932.062/0001-50"""
    d = re.sub(r"\D", "", cnpj or "")
    if len(d) != 14:
        return cnpj
    return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"


def identificar_estabelecimento(cnpj: str) -> str:
    """0001 -> 'Matriz'; demais -> 'Filial XXXX'"""
    d = re.sub(r"\D", "", cnpj or "")
    if len(d) != 14:
        return "?"
    seq = d[8:12]
    return "Matriz" if seq == "0001" else f"Filial {seq}"


# ============================================================================
# GERENCIADOR DE CERTIFICADOS
# ============================================================================

def extrair_metadados_pfx(pfx_bytes: bytes, password: str) -> dict:
    """Carrega um PFX e extrai metadados do certificado.

    Retorna um dict com sempre as chaves: ok, erro, e (se ok) tipo, cnpj,
    razao_social, estabelecimento, valid_from, valid_to, expired, dias_restantes.
    """
    try:
        senha_bytes = password.encode() if password else None
        _priv, cert, _add = pkcs12.load_key_and_certificates(pfx_bytes, senha_bytes)
        if cert is None:
            return {"ok": False, "erro": "PFX sem certificado válido"}

        # Common Name (formato ICP-Brasil: "RAZAO SOCIAL:DOCUMENTO")
        cn_attrs = cert.subject.get_attributes_for_oid(NameOID.COMMON_NAME)
        cn = cn_attrs[0].value if cn_attrs else ""

        razao_social = cn
        documento = ""
        if ":" in cn:
            partes = cn.rsplit(":", 1)
            razao_social = partes[0].strip()
            documento_raw = partes[1].strip()
            documento = re.sub(r"\D", "", documento_raw)

        # Tipo: e-CNPJ (14d) vs e-CPF (11d) vs outro
        if len(documento) == 14:
            tipo = "e-CNPJ"
            cnpj = documento
            cpf = ""
            estabelecimento = identificar_estabelecimento(cnpj)
        elif len(documento) == 11:
            tipo = "e-CPF"
            cnpj = ""
            cpf = documento
            estabelecimento = "-"
        else:
            tipo = "Desconhecido"
            cnpj = ""
            cpf = ""
            estabelecimento = "-"

        # Validade
        valid_from = getattr(cert, "not_valid_before_utc", None) or cert.not_valid_before
        valid_to = getattr(cert, "not_valid_after_utc", None) or cert.not_valid_after
        if valid_to.tzinfo is None:
            valid_to_aware = valid_to.replace(tzinfo=timezone.utc)
        else:
            valid_to_aware = valid_to
        agora = datetime.now(timezone.utc)
        dias_restantes = (valid_to_aware - agora).days
        expired = dias_restantes < 0

        return {
            "ok": True,
            "tipo": tipo,
            "cn": cn,
            "razao_social": razao_social,
            "cnpj": cnpj,
            "cpf": cpf,
            "estabelecimento": estabelecimento,
            "valid_from": valid_from,
            "valid_to": valid_to,
            "expired": expired,
            "dias_restantes": dias_restantes,
            "issuer": cert.issuer.rfc4514_string(),
        }

    except ValueError as e:
        if "Invalid password" in str(e) or "mac verify" in str(e).lower():
            return {"ok": False, "erro": "senha incorreta"}
        return {"ok": False, "erro": f"PFX inválido: {e}"}
    except Exception as e:
        return {"ok": False, "erro": f"{type(e).__name__}: {e}"}


def varrer_pasta_pfx(pasta: Path) -> list:
    """Le todos os .pfx/.p12 de uma pasta. Retorna [(filename, bytes), ...]"""
    if not pasta.exists() or not pasta.is_dir():
        return []
    arquivos = []
    for ext in ("*.pfx", "*.p12", "*.PFX", "*.P12"):
        for p in pasta.glob(ext):
            arquivos.append((p.name, p.read_bytes()))
    return arquivos


# ============================================================================
# PARSER DA NFS-e (XML)
# ============================================================================

def parse_nfse_xml(xml_bytes: bytes) -> dict:
    root = ET.fromstring(xml_bytes)

    def find(elt, path):
        n = elt.find(path, NS)
        if n is None:
            n = elt.find(re.sub(r"nfse:", "", path))
        return n

    def text(elt, path, default=""):
        n = find(elt, path)
        return (n.text or default).strip() if n is not None else default

    chave = ""
    inf = find(root, ".//nfse:infNFSe")
    if inf is not None:
        chave = (inf.get("Id") or "").replace("NFSe", "")
    if not chave:
        chave = text(root, ".//nfse:chNFSe")

    numero = text(root, ".//nfse:nNFSe")
    data_emissao_raw = text(root, ".//nfse:dhEmi") or text(root, ".//nfse:DataEmissao")

    try:
        dt = datetime.fromisoformat(data_emissao_raw.replace("Z", "+00:00"))
        data_emissao = dt.strftime("%Y-%m-%d")
        competencia = dt.strftime("%Y-%m")
    except Exception:
        data_emissao = data_emissao_raw[:10] if data_emissao_raw else "0000-00-00"
        competencia = data_emissao[:7] if data_emissao else "0000-00"

    cnpj_emitente = (
        text(root, ".//nfse:emit/nfse:CNPJ")
        or text(root, ".//nfse:prest/nfse:CNPJ")
        or text(root, ".//nfse:CNPJEmit")
    )
    nome_emitente = (
        text(root, ".//nfse:emit/nfse:xNome")
        or text(root, ".//nfse:prest/nfse:xNome")
        or text(root, ".//nfse:xNomeEmit")
    )

    valor_str = text(root, ".//nfse:vServ") or text(root, ".//nfse:vLiq") or "0"
    try:
        valor = float(valor_str.replace(",", "."))
    except ValueError:
        valor = 0.0

    cnpj_tomador = (
        text(root, ".//nfse:toma/nfse:CNPJ")
        or text(root, ".//nfse:tomador/nfse:CNPJ")
    )

    return {
        "chave": chave,
        "numero": numero,
        "data_emissao": data_emissao,
        "competencia": competencia,
        "cnpj_emitente": cnpj_emitente,
        "nome_emitente": nome_emitente,
        "valor_servico": valor,
        "cnpj_tomador": cnpj_tomador,
    }


def montar_nome_arquivo(meta: dict) -> str:
    data = meta["data_emissao"]
    numero = (meta["numero"] or "0").zfill(8)
    cnpj8 = (meta["cnpj_emitente"] or "00000000")[:8]
    nome = slugify(meta["nome_emitente"], max_len=25)
    return f"{data}_NFSE-{numero}_{cnpj8}_{nome}"


def decodificar_xml_recebido(xml_data) -> bytes:
    """Decodifica o conteudo do campo ArquivoXml do ADN.

    O ADN devolve o XML em base64. Internamente, pode estar:
    - XML puro (eventos costumam vir assim)
    - GZIP comprimido (NFS-e costumam vir comprimidas)
    - ZIP (formato menos comum, mas suportado)

    Retorna sempre bytes do XML descomprimido pronto pra parse.
    """
    if not xml_data:
        return b""

    # Passo 1: base64 -> bytes brutos
    try:
        if isinstance(xml_data, str):
            raw = base64.b64decode(xml_data)
        else:
            raw = xml_data
    except Exception:
        raw = xml_data.encode("utf-8") if isinstance(xml_data, str) else xml_data

    if not raw:
        return b""

    # Passo 2: detectar formato pelos magic bytes
    # GZIP: 1F 8B
    if raw[:2] == b"\x1f\x8b":
        try:
            return gzip.decompress(raw)
        except Exception as e:
            # Se falhar, tenta GzipFile (algumas variacoes nao decodificam com decompress)
            try:
                return gzip.GzipFile(fileobj=io.BytesIO(raw)).read()
            except Exception:
                pass

    # ZIP: 50 4B 03 04 (PK..)
    if raw[:2] == b"PK":
        try:
            with zipfile.ZipFile(io.BytesIO(raw)) as zf:
                for name in zf.namelist():
                    if name.lower().endswith(".xml"):
                        return zf.read(name)
                # Se nao achou .xml, retorna o primeiro arquivo
                if zf.namelist():
                    return zf.read(zf.namelist()[0])
        except Exception:
            pass

    # Caso ja seja XML puro (comeca com < ou BOM)
    if raw[:1] == b"<" or raw[:3] == b"\xef\xbb\xbf":
        return raw

    # Ultimo recurso: devolve o raw mesmo (parse vai falhar e logar)
    return raw


# ============================================================================
# CLIENTE ADN
# ============================================================================

class NFSeNacionalClient:
    def __init__(self, pfx_bytes: bytes, pfx_password: str, homologacao: bool = False):
        self.adn_base = ADN_HOMOLOG if homologacao else ADN_PROD
        self.nfse_base = NFSE_HOMOLOG if homologacao else NFSE_PROD

        # Persiste o PFX num arquivo temporario cross-platform.
        # No Windows, isso vira algo como C:\Users\Pedro\AppData\Local\Temp\
        # No Linux/macOS, /tmp/. requests_pkcs12 trabalha melhor com arquivo em disco.
        self._pfx_path = Path(tempfile.gettempdir()) / "cert_nfse_rezende.pfx"
        self._pfx_path.write_bytes(pfx_bytes)

        self.session = requests.Session()
        adapter = Pkcs12Adapter(
            pkcs12_filename=str(self._pfx_path),
            pkcs12_password=pfx_password,
        )
        self.session.mount(self.adn_base, adapter)
        self.session.mount(self.nfse_base, adapter)
        self.session.headers.update({"Accept": "application/json"})

    def baixar_lote(self, ultimo_nsu: int = 0) -> dict:
        """Baixa um lote de ate 50 NFS-e via ADN.

        Endpoint oficial (contribuintes): /contribuintes/DFe/{ultNSU}
        Documentacao: https://adn.nfse.gov.br/contribuintes/docs/index.html
        """
        url = f"{self.adn_base}/contribuintes/DFe/{ultimo_nsu}"
        r = self.session.get(url, timeout=60)

        # Log detalhado pra debug
        if r.status_code != 200:
            raise requests.HTTPError(
                f"HTTP {r.status_code} em {url}\n"
                f"Headers: {dict(r.headers)}\n"
                f"Body (primeiros 500 chars): {r.text[:500]}",
                response=r,
            )

        ctype = r.headers.get("Content-Type", "")
        if "json" in ctype:
            return r.json()
        return self._parse_envelope_xml(r.content)

    @staticmethod
    def _parse_envelope_xml(xml_bytes: bytes) -> dict:
        root = ET.fromstring(xml_bytes)
        docs = []
        for doc in root.iter():
            if doc.tag.endswith("docZip") or doc.tag.endswith("ArquivoXML"):
                docs.append({
                    "NSU": doc.get("NSU") or "",
                    "ChaveAcesso": doc.get("chNFSe") or "",
                    "ArquivoXML": doc.text or "",
                })
        return {"LoteDFe": docs}

    def baixar_danfse(self, chave_acesso: str) -> Optional[bytes]:
        """Baixa o DANFSe (PDF) oficial via endpoint ADN.

        Endpoint oficial: /danfse/{chave}
        Documentacao: https://adn.nfse.gov.br/danfse/docs/index.html
        """
        candidatas = [
            f"{self.adn_base}/danfse/{chave_acesso}",
            f"{self.adn_base}/danfse/NFSe/{chave_acesso}",
            f"{self.adn_base}/contribuintes/DANFSe/{chave_acesso}",
        ]
        for url in candidatas:
            try:
                r = self.session.get(url, timeout=30,
                                     headers={"Accept": "application/pdf"})
                if r.status_code == 200 and r.content[:4] == b"%PDF":
                    return r.content
            except requests.RequestException:
                continue
        return None


# ============================================================================
# ORQUESTRADOR DO DOWNLOAD
# ============================================================================

def executar_download(client: NFSeNacionalClient, pasta_saida: Path,
                      nsu_inicial: int, baixar_pdf: bool,
                      max_lotes: int, progress_cb) -> pd.DataFrame:
    pasta_pdf = pasta_saida / "PDFs"
    pasta_xml = pasta_saida / "XMLs"
    pasta_eventos = pasta_saida / "Eventos"
    pasta_pdf.mkdir(parents=True, exist_ok=True)
    pasta_xml.mkdir(parents=True, exist_ok=True)
    pasta_eventos.mkdir(parents=True, exist_ok=True)

    manifesto = []
    nsu = nsu_inicial
    lotes = 0

    while lotes < max_lotes:
        url_chamada = f"{client.adn_base}/contribuintes/DFe/{nsu}"
        progress_cb(f"Buscando lote: GET {url_chamada}")
        try:
            lote = client.baixar_lote(nsu)
        except requests.HTTPError as e:
            progress_cb(f"ERRO HTTP: {e}")
            break
        except Exception as e:
            progress_cb(f"ERRO {type(e).__name__}: {e}")
            break

        documentos = (
            lote.get("LoteDFe")
            or lote.get("loteDFe")
            or lote.get("DFe")
            or lote.get("documentos")
            or lote.get("Documentos")
            or lote.get("docZip")
            or []
        )

        if isinstance(lote, dict):
            keys = list(lote.keys())[:10]
            progress_cb(f"  Response keys: {keys}")

        if not documentos:
            if isinstance(lote, dict) and lote:
                import json as _json
                preview = _json.dumps(lote, ensure_ascii=False, default=str)[:400]
                progress_cb(f"  Lote sem documentos. Payload: {preview}")
            else:
                progress_cb("Sem mais documentos. Pull completo.")
            break

        progress_cb(f"  {len(documentos)} documento(s) no lote.")

        nsu_max_lote = nsu
        nfse_count = 0
        evento_count = 0
        falhas_count = 0

        for doc in documentos:
            nsu_doc = int(doc.get("NSU") or doc.get("nsu") or (nsu + 1))
            nsu_max_lote = max(nsu_max_lote, nsu_doc)

            xml_b64 = (
                doc.get("ArquivoXml")
                or doc.get("ArquivoXML")
                or doc.get("arquivoXml")
                or doc.get("arquivoXML")
                or doc.get("XMLNFSe")
                or doc.get("xmlNFSe")
                or doc.get("docZip")
                or doc.get("xml")
                or ""
            )
            tipo_doc = doc.get("TipoDocumento") or doc.get("tipoDocumento") or ""
            tipo_evento = doc.get("TipoEvento") or doc.get("tipoEvento") or ""
            chave_doc = doc.get("ChaveAcesso") or doc.get("chaveAcesso") or ""

            if not xml_b64:
                progress_cb(f"  NSU {nsu_doc}: sem campo de XML")
                continue

            # Decodifica (base64 + descompactacao gzip/zip se necessario)
            xml_bytes = decodificar_xml_recebido(xml_b64)
            if not xml_bytes:
                progress_cb(f"  NSU {nsu_doc}: falha ao decodificar")
                continue

            # Eventos
            if tipo_evento or "EVENTO" in str(tipo_doc).upper():
                comp_evento = datetime.now().strftime("%Y-%m")
                sub_ev = pasta_eventos / comp_evento
                sub_ev.mkdir(parents=True, exist_ok=True)
                ev_path = sub_ev / f"EVENTO_NSU-{nsu_doc:08d}_{chave_doc[:20] or 'sem-chave'}.xml"
                ev_path.write_bytes(xml_bytes)
                evento_count += 1
                manifesto.append({
                    "NSU": nsu_doc,
                    "Chave": chave_doc,
                    "Numero": "",
                    "Data Emissão": "",
                    "Competência": comp_evento,
                    "CNPJ Emitente": "",
                    "Razão Social": f"[EVENTO {tipo_evento}]",
                    "Valor (R$)": 0.0,
                    "CNPJ Tomador": "",
                    "Arquivo XML": str(ev_path.relative_to(pasta_saida)),
                    "PDF": "-",
                })
                progress_cb(f"  NSU {nsu_doc} | EVENTO {tipo_evento} | salvo")
                continue

            # NFS-e: tenta o parse completo
            try:
                meta = parse_nfse_xml(xml_bytes)
            except ET.ParseError as e:
                # Salva o XML cru mesmo assim na pasta "_falhas" para diagnostico
                pasta_falhas = pasta_saida / "_falhas_parse"
                pasta_falhas.mkdir(exist_ok=True)
                falha_path = pasta_falhas / f"NSU-{nsu_doc:08d}.xml"
                falha_path.write_bytes(xml_bytes)
                # Tambem dumpa um sample binario do primeiro caso pra debug
                if falhas_count == 0:
                    debug_path = pasta_falhas / f"NSU-{nsu_doc:08d}.raw.bin"
                    try:
                        debug_path.write_bytes(base64.b64decode(xml_b64))
                    except Exception:
                        pass
                    progress_cb(f"  NSU {nsu_doc}: parse falhou ({e}). XML salvo em _falhas_parse/")
                falhas_count += 1
                continue

            nome_base = montar_nome_arquivo(meta)
            comp = meta["competencia"]

            sub_xml = pasta_xml / comp
            sub_pdf = pasta_pdf / comp
            sub_xml.mkdir(parents=True, exist_ok=True)
            sub_pdf.mkdir(parents=True, exist_ok=True)

            xml_path = sub_xml / f"{nome_base}.xml"
            xml_path.write_bytes(xml_bytes)

            pdf_status = "-"
            if baixar_pdf and meta["chave"]:
                pdf_bytes = client.baixar_danfse(meta["chave"])
                if pdf_bytes:
                    pdf_path = sub_pdf / f"{nome_base}.pdf"
                    pdf_path.write_bytes(pdf_bytes)
                    pdf_status = "OK"
                else:
                    pdf_status = "falha"
                time.sleep(0.3)

            nfse_count += 1
            manifesto.append({
                "NSU": nsu_doc,
                "Chave": meta["chave"],
                "Numero": meta["numero"],
                "Data Emissão": meta["data_emissao"],
                "Competência": meta["competencia"],
                "CNPJ Emitente": meta["cnpj_emitente"],
                "Razão Social": meta["nome_emitente"],
                "Valor (R$)": meta["valor_servico"],
                "CNPJ Tomador": meta["cnpj_tomador"],
                "Arquivo XML": str(xml_path.relative_to(pasta_saida)),
                "PDF": pdf_status,
            })

            progress_cb(f"  NSU {nsu_doc} | {meta['data_emissao']} | "
                       f"{meta['nome_emitente'][:35]} | R$ {meta['valor_servico']:.2f}")

        nsu = nsu_max_lote
        lotes += 1
        progress_cb(f"  Lote {lotes}: {nfse_count} NFS-e + {evento_count} eventos + "
                    f"{falhas_count} falhas. Proximo NSU base: {nsu}")

        time.sleep(1.5)

        if len(documentos) < 50:
            progress_cb("Ultimo lote (menos de 50 documentos).")
            break

    return pd.DataFrame(manifesto)


def gerar_manifesto_xlsx(df: pd.DataFrame, caminho: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "NFS-e Recebidas"

    if df.empty:
        ws["A1"] = "Nenhuma NFS-e baixada."
        wb.save(caminho)
        return

    cols = list(df.columns)
    for c, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="F7931E")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r, row in enumerate(df.itertuples(index=False), 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    larguras = {
        "NSU": 10, "Chave": 50, "Numero": 12, "Data Emissão": 14,
        "Competência": 13, "CNPJ Emitente": 18, "Razão Social": 40,
        "Valor (R$)": 14, "CNPJ Tomador": 18, "Arquivo XML": 60, "PDF": 10,
    }
    for c, col in enumerate(cols, 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = larguras.get(col, 18)

    ws2 = wb.create_sheet("Resumo por Competência")
    resumo = df.groupby("Competência").agg(
        Qtd_Notas=("Chave", "count"),
        Valor_Total=("Valor (R$)", "sum"),
    ).reset_index()
    for c, col in enumerate(resumo.columns, 1):
        cell = ws2.cell(row=1, column=c, value=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="F7931E")
    for r, row in enumerate(resumo.itertuples(index=False), 2):
        for c, val in enumerate(row, 1):
            ws2.cell(row=r, column=c, value=val)
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 18

    wb.save(caminho)


# ============================================================================
# INTERFACE STREAMLIT
# ============================================================================

def render_card_certificado(meta: dict) -> str:
    """HTML de um card de certificado, com badges e cores."""
    if meta["expired"]:
        badge_val = f'<span style="background:#7A1F1F;color:#FFB3B3;padding:2px 8px;border-radius:4px;font-size:11px;">VENCIDO há {abs(meta["dias_restantes"])}d</span>'
    elif meta["dias_restantes"] < 30:
        badge_val = f'<span style="background:#7A5F1F;color:#FFE0A3;padding:2px 8px;border-radius:4px;font-size:11px;">vence em {meta["dias_restantes"]}d</span>'
    else:
        badge_val = f'<span style="background:#1F5F2F;color:#A3FFB8;padding:2px 8px;border-radius:4px;font-size:11px;">{meta["dias_restantes"]}d restantes</span>'

    if meta["tipo"] == "e-CPF":
        badge_tipo = '<span style="background:#7A1F1F;color:#FFB3B3;padding:2px 8px;border-radius:4px;font-size:11px;">e-CPF (nao serve)</span>'
    elif meta["estabelecimento"] == "Matriz":
        badge_tipo = '<span style="background:#F7931E;color:#0A0A0A;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:700;">MATRIZ</span>'
    else:
        badge_tipo = f'<span style="background:#3A6FAA;color:#E0F0FF;padding:2px 8px;border-radius:4px;font-size:11px;">{meta["estabelecimento"].upper()}</span>'

    rezende_match = ""
    if meta["cnpj"].startswith(CNPJ_BASE_REZENDE):
        rezende_match = '<span style="background:#1F5F2F;color:#A3FFB8;padding:2px 8px;border-radius:4px;font-size:11px;">REZENDE</span>'

    doc = formatar_cnpj(meta["cnpj"]) if meta["cnpj"] else meta["cpf"]
    val = meta["valid_to"].strftime("%d/%m/%Y") if hasattr(meta["valid_to"], "strftime") else str(meta["valid_to"])

    return f"""
    <div style="background:#1A1A1A;border:1px solid #2A2A2A;border-radius:8px;padding:12px;margin-bottom:8px;">
        <div style="font-weight:700;color:#FFF;font-size:14px;margin-bottom:4px;">
            {meta['razao_social']}
        </div>
        <div style="font-family:'DM Mono',monospace;font-size:12px;color:#BBB;margin-bottom:6px;">
            {doc} &middot; valido ate {val}
        </div>
        <div>{badge_tipo} {badge_val} {rezende_match}</div>
        <div style="font-size:10px;color:#666;margin-top:6px;">arquivo: {meta['filename']}</div>
    </div>
    """


def main():
    st.set_page_config(
        page_title="NFS-e Recebidas - Rezende",
        page_icon=":inbox_tray:",
        layout="wide",
    )

    st.markdown(f"""
    <style>
        .stApp {{ background-color: {BG}; color: #EEE; }}
        h1, h2, h3 {{ color: {ORANGE}; }}
        .stButton>button {{
            background-color: {ORANGE}; color: {BG};
            font-weight: 700; border: none;
        }}
        .stButton>button:hover {{ background-color: #FFA640; color: {BG}; }}
        .stTextInput input, .stNumberInput input, .stTextArea textarea {{
            background-color: #1A1A1A; color: #FFF; border: 1px solid #333;
        }}
        .log-box {{
            background-color: #0F0F0F; border: 1px solid #2A2A2A;
            border-radius: 6px; padding: 12px;
            font-family: 'DM Mono', monospace; font-size: 12px;
            max-height: 400px; overflow-y: auto; color: #DDD;
            white-space: pre-wrap;
        }}
    </style>
    """, unsafe_allow_html=True)

    st.title("NFS-e Recebidas - Pull em Lote")
    st.caption("Rezende Construcao e Manutencao - API oficial ADN do Sistema Nacional NFS-e")

    # Estado da sessão
    if "certs_identificados" not in st.session_state:
        st.session_state.certs_identificados = []
    if "cert_falhas" not in st.session_state:
        st.session_state.cert_falhas = []

    # ------------------------------------------------------------------
    # SIDEBAR
    # ------------------------------------------------------------------
    with st.sidebar:
        st.header("Certificado Digital")

        modo_cert = st.radio(
            "Como fornecer os certificados?",
            ["Upload (anexar arquivos)", "Pasta local (varrer diretorio)"],
            index=0,
        )

        pfxs_disponiveis = []

        if modo_cert == "Upload (anexar arquivos)":
            uploaded = st.file_uploader(
                "Anexe um ou mais certificados (.pfx / .p12)",
                type=["pfx", "p12"],
                accept_multiple_files=True,
                help="Pode anexar todos os seus 5 certificados de uma vez - "
                     "o app identifica qual e a Matriz e qual e a Filial.",
            )
            if uploaded:
                for f in uploaded:
                    pfxs_disponiveis.append((f.name, f.getvalue()))
        else:
            pasta_certs_str = st.text_input(
                "Caminho da pasta com os PFX",
                value=str(Path.home() / "Documentos" / "Certificados"),
                help="Ex.: C:\\Users\\Pedro\\Documentos\\Certificados",
            )
            if pasta_certs_str:
                arquivos = varrer_pasta_pfx(Path(pasta_certs_str))
                pfxs_disponiveis = arquivos
                if arquivos:
                    st.caption(f"{len(arquivos)} arquivo(s) encontrado(s) na pasta.")
                else:
                    st.caption("Nenhum .pfx/.p12 encontrado nessa pasta.")

        password = st.text_input(
            "Senha (sera tentada em todos os PFX)",
            type="password",
            help="Se cada certificado tem senha diferente, identifique-os um por vez.",
        )

        if st.button("Identificar certificados", use_container_width=True):
            if not pfxs_disponiveis:
                st.error("Nenhum arquivo PFX disponivel.")
            elif not password:
                st.error("Informe a senha.")
            else:
                identificados = []
                falhas = []
                for filename, pfx_bytes in pfxs_disponiveis:
                    meta = extrair_metadados_pfx(pfx_bytes, password)
                    if meta["ok"]:
                        meta["filename"] = filename
                        meta["bytes"] = pfx_bytes
                        meta["password"] = password
                        identificados.append(meta)
                    else:
                        falhas.append({"filename": filename, "erro": meta["erro"]})
                st.session_state.certs_identificados = identificados
                st.session_state.cert_falhas = falhas

        st.divider()
        st.header("Parametros do Pull")

        ambiente = st.selectbox(
            "Ambiente",
            ["Producao", "Homologacao (testes)"],
            index=0,
        )
        nsu_inicial = st.number_input(
            "NSU inicial", min_value=0, value=0, step=1,
            help="0 na 1a execucao. Depois, o maior NSU baixado.",
        )
        max_lotes = st.number_input(
            "Max. lotes (50 docs cada)",
            min_value=1, max_value=200, value=50, step=1,
        )
        baixar_pdf = st.checkbox("Baixar PDF (DANFSe) alem do XML", value=True)
        pasta_saida_str = st.text_input(
            "Pasta de saida", value="./notas_fiscais_recebidas",
        )

    # ------------------------------------------------------------------
    # MAIN
    # ------------------------------------------------------------------
    col_esq, col_dir = st.columns([1, 1])

    with col_esq:
        st.subheader("Certificados identificados")

        if not st.session_state.certs_identificados and not st.session_state.cert_falhas:
            st.info("Forneca os certificados e a senha no painel lateral, "
                    "depois clique em **Identificar certificados**.")
        else:
            for meta in st.session_state.certs_identificados:
                st.markdown(
                    render_card_certificado(meta),
                    unsafe_allow_html=True,
                )

            if st.session_state.cert_falhas:
                with st.expander(
                    f"{len(st.session_state.cert_falhas)} certificado(s) "
                    "nao puderam ser lidos"
                ):
                    for f in st.session_state.cert_falhas:
                        st.markdown(f"- **{f['filename']}** -- {f['erro']}")
                    st.caption("Provavelmente senha diferente. Identifique-os "
                               "separadamente com a senha correta.")

    with col_dir:
        st.subheader("Execucao")

        if not st.session_state.certs_identificados:
            st.warning("Identifique pelo menos um certificado antes de executar.")
            return

        elegíveis = [
            c for c in st.session_state.certs_identificados
            if c["tipo"] == "e-CNPJ" and not c["expired"]
        ]

        if not elegíveis:
            st.error(
                "Nenhum certificado e-CNPJ valido disponivel. "
                "Verifique se os certificados nao estao vencidos e se nao sao e-CPF."
            )
            return

        opcoes = {
            f"{c['razao_social'][:40]} -- {formatar_cnpj(c['cnpj'])} ({c['estabelecimento']})": c
            for c in elegíveis
        }

        chave_padrao = next(
            (k for k, v in opcoes.items()
             if v["cnpj"] == CNPJ_REZENDE_MATRIZ),
            list(opcoes.keys())[0],
        )

        escolha = st.selectbox(
            "Certificado para a consulta:",
            options=list(opcoes.keys()),
            index=list(opcoes.keys()).index(chave_padrao),
        )
        cert_selecionado = opcoes[escolha]

        st.markdown(
            f"**CNPJ do certificado:** `{formatar_cnpj(cert_selecionado['cnpj'])}`  \n"
            f"**Tipo:** {cert_selecionado['estabelecimento']}  \n"
            f"**Validade:** ate {cert_selecionado['valid_to'].strftime('%d/%m/%Y')}"
        )

        if not cert_selecionado["cnpj"].startswith(CNPJ_BASE_REZENDE):
            st.warning(
                "Este certificado **nao e da Rezende** "
                f"(raiz {CNPJ_BASE_REZENDE}). Voce so vai receber notas "
                "enderecadas ao CNPJ deste certificado."
            )

        st.divider()

        if st.button("Iniciar download em lote", use_container_width=True):
            pasta_saida = Path(pasta_saida_str).expanduser().resolve()
            pasta_saida.mkdir(parents=True, exist_ok=True)

            log_placeholder = st.empty()
            logs = []

            def progress_cb(msg: str):
                logs.append(msg)
                log_placeholder.markdown(
                    f"<div class='log-box'>{'<br>'.join(logs[-200:])}</div>",
                    unsafe_allow_html=True,
                )

            with st.spinner("Conectando ao ADN e baixando notas..."):
                try:
                    client = NFSeNacionalClient(
                        pfx_bytes=cert_selecionado["bytes"],
                        pfx_password=cert_selecionado["password"],
                        homologacao=(ambiente == "Homologacao (testes)"),
                    )
                    progress_cb(f"Certificado: {cert_selecionado['razao_social']} "
                               f"({cert_selecionado['estabelecimento']})")
                    progress_cb(f"Ambiente: {ambiente}")
                    progress_cb(f"Saida: {pasta_saida}")

                    df = executar_download(
                        client=client,
                        pasta_saida=pasta_saida,
                        nsu_inicial=int(nsu_inicial),
                        baixar_pdf=baixar_pdf,
                        max_lotes=int(max_lotes),
                        progress_cb=progress_cb,
                    )

                    manifesto_path = pasta_saida / "manifesto.xlsx"
                    gerar_manifesto_xlsx(df, manifesto_path)
                    progress_cb(f"Manifesto: {manifesto_path}")

                    st.success(f"{len(df)} NFS-e processadas.")

                    if not df.empty:
                        c1, c2, c3 = st.columns(3)
                        c1.metric("Notas baixadas", len(df))
                        c2.metric("Valor total", f"R$ {df['Valor (R$)'].sum():,.2f}")
                        c3.metric("Maior NSU", int(df["NSU"].max()))
                        st.dataframe(df, use_container_width=True, height=400)

                        with open(manifesto_path, "rb") as f:
                            st.download_button(
                                "Baixar manifesto.xlsx",
                                data=f,
                                file_name="manifesto_nfse_recebidas.xlsx",
                                mime="application/vnd.openxmlformats-officedocument."
                                     "spreadsheetml.sheet",
                            )

                except Exception as e:
                    st.error(f"{type(e).__name__}: {e}")
                    import traceback
                    st.code(traceback.format_exc())


if __name__ == "__main__":
    main()